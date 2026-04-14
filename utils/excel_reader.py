"""Excel Reader for Robot Framework using openpyxl"""
import openpyxl

def read_excel_data(file_path):
    """Reads Excel file and returns dictionary"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    data = {}
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    
    for i, header in enumerate(headers):
        if header and i < len(values):
            data[str(header)] = str(values[i]) if values[i] is not None else ''
    
    wb.close()
    return data

def read_excel_sheet(file_path, sheet_name):
    """Reads specific sheet from Excel file and returns dictionary"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    data = {}
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    
    for i, header in enumerate(headers):
        if header and i < len(values):
            data[str(header)] = str(values[i]) if values[i] is not None else ''
    
    wb.close()
    return data

def read_excel_row(file_path, header_row, data_row):
    """Reads specific row with its header row from Excel file"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    data = {}
    headers = [cell.value for cell in ws[header_row]]
    values = [cell.value for cell in ws[data_row]]
    
    for i, header in enumerate(headers):
        if header and i < len(values):
            data[str(header)] = str(values[i]) if values[i] is not None else ''
    
    wb.close()
    return data
