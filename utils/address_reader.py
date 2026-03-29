import openpyxl
from datetime import datetime

def read_address_sections(file_path):
    """Read all 3 address sections from Excel.
    Row 3  = RESIDENTIAL ADDRESS
    Row 7  = PERMANENT ADDRESS
    Row 11 = BUSINESS ADDRESS
    Columns:
        A=Action, B=Customer Type, C=Status, D=Living Since,
        E=District, F=Area, G=Block, H=Street, I=Building,
        J=Unit Type, K=Unit No, L=Floor
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    addresses = {
        'RESIDENTIAL': {},
        'PERMANENT':   {},
        'BUSINESS':    {},
    }

    section_map = {
        'RESIDENTIAL': 3,
        'PERMANENT':   7,
        'BUSINESS':   11,
    }

    # Address type mapping based on section
    addr_type_map = {
        'RESIDENTIAL': 'RESIDENTIAL ADDRESS',
        'PERMANENT':   'PERMANENT ADDRESS',
        'BUSINESS':    'BUSINESS ADDRESS',
    }

    def format_date(date_value):
        """Convert date to DD-MM-YYYY format"""
        if not date_value:
            return ''
        
        # If it's a datetime object from Excel
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%d-%m-%Y')
        
        date_str = str(date_value).strip()
        
        if not date_str or date_str.lower() == 'none':
            return ''
        
        # If already in DD-MM-YYYY format, return as is
        if '-' in date_str and len(date_str.split('-')) == 3:
            parts = date_str.split('-')
            if len(parts[0]) == 2:  # DD-MM-YYYY
                return date_str
        
        # Try to parse common date formats
        for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S']:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime('%d-%m-%Y')
            except ValueError:
                continue
        
        # If parsing fails, return original
        return date_str

    for key, row in section_map.items():
        action = ws.cell(row=row, column=1).value
        if action and str(action).strip().lower() != 'skip':
            living_since_raw = ws.cell(row=row, column=4).value
            living_since_formatted = format_date(living_since_raw)
            
            addresses[key] = {
                'Action':        str(ws.cell(row=row, column=1).value or '').strip(),
                'Customer_Type': str(ws.cell(row=row, column=2).value or '').strip(),
                'Address_Type':  addr_type_map[key],
                'Status':        str(ws.cell(row=row, column=3).value or '').strip(),
                'ID_Type':       'CIVIL ID Number',
                'ID_Number':     '',
                'Living_Since':  living_since_formatted,
                'District':      str(ws.cell(row=row, column=5).value or '').strip(),
                'Area':          str(ws.cell(row=row, column=6).value or '').strip(),
                'Block':         str(ws.cell(row=row, column=7).value or '').strip(),
                'Street':        str(ws.cell(row=row, column=8).value or '').strip(),
                'Building':      str(ws.cell(row=row, column=9).value or '').strip(),
                'Unit_Type':     str(ws.cell(row=row, column=10).value or '').strip(),
                'Unit_No':       str(ws.cell(row=row, column=11).value or '').strip(),
                'Floor':         str(ws.cell(row=row, column=12).value or '').strip(),
                'Address_Line':  '',
            }

    return addresses
