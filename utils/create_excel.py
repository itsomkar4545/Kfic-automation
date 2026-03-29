from openpyxl import Workbook
import os

# Create data directory if it doesn't exist
data_dir = "../../data"
os.makedirs(data_dir, exist_ok=True)

# Create customer_search.xlsx
wb_customer = Workbook()
ws_customer = wb_customer.active
ws_customer.title = "Sheet1"
ws_customer['A1'] = 'Customer ID'
ws_customer['A2'] = '12345678901'
wb_customer.save(f"{data_dir}/customer_search.xlsx")

# Create change_address.xlsx with multiple sheets
wb_address = Workbook()

# Remove default sheet
wb_address.remove(wb_address.active)

# RESIDENTIAL sheet
ws_res = wb_address.create_sheet("RESIDENTIAL")
headers = ['Action', 'Address_Type', 'Status', 'Area', 'Block', 'Street', 'Building', 'Unit_Type', 'Unit_No', 'Floor', 'Address_Line', 'Customer_Type', 'District', 'Living_Since']
for i, header in enumerate(headers, 1):
    ws_res.cell(row=1, column=i, value=header)
data = ['Edit', 'RESIDENTIAL ADDRESS', 'Active', 'SALMIYA', '1', '10', 'Building A', 'Apartment', '101', '1', 'Sample Address Line', 'Main Applicant', 'HAWALLI', '01/01/2020']
for i, value in enumerate(data, 1):
    ws_res.cell(row=2, column=i, value=value)

# PERMANENT sheet
ws_perm = wb_address.create_sheet("PERMANENT")
for i, header in enumerate(headers, 1):
    ws_perm.cell(row=1, column=i, value=header)
data = ['Create', 'PERMANENT ADDRESS', 'Active', 'JABRIYA', '2', '20', 'Building B', 'Villa', '201', '2', 'Permanent Address Line', 'Main Applicant', 'HAWALLI', '01/01/2019']
for i, value in enumerate(data, 1):
    ws_perm.cell(row=2, column=i, value=value)

# BUSINESS sheet
ws_bus = wb_address.create_sheet("BUSINESS")
for i, header in enumerate(headers, 1):
    ws_bus.cell(row=1, column=i, value=header)
data = ['Skip', 'BUSINESS ADDRESS', 'Active', 'AHMADI', '3', '30', 'Building C', 'Apartment', '301', '3', 'Business Address Line', 'Main Applicant', 'AHMADI', '01/01/2021']
for i, value in enumerate(data, 1):
    ws_bus.cell(row=2, column=i, value=value)

wb_address.save(f"{data_dir}/change_address.xlsx")

print("Excel files created successfully!")