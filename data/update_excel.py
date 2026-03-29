import openpyxl

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Row 1 - Headers
ws['A1'] = 'ID_Type'
ws['B1'] = 'ID_Number'
ws['C1'] = 'Customer_Name'
ws['D1'] = 'Nationality'
ws['E1'] = 'Email'
ws['F1'] = 'Mobile_Number'
ws['G1'] = 'Product'
ws['H1'] = 'SubProduct'

# Row 2 - Data
ws['A2'] = 'Civil ID Number'
ws['B2'] = '341164000000'
ws['C2'] = 'ADIL RASHID'
ws['D2'] = 'BANGLADESH'
ws['E2'] = 'ADILR74@GMAIL.COM'
ws['F2'] = '81234567'
ws['G2'] = 'VEHICLE LOAN RETAIL'
ws['H2'] = '01-NEW CAR – AGENCY'

# Save
wb.save('retail_flow_input.xlsx')
print("Excel file created successfully!")
