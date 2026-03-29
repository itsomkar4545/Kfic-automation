"""Generate input Excel files for all Service Request types"""
from openpyxl import Workbook
import os

SERVICE_REQUESTS = [
    "employment_request",
    "personal_details",
    "refund_maker",
    "valuation_sr",
    "add_collateral_sr",
    "change_address",
    "cinet_cancel",
    "debt_ack_cancel",
    "repayment_mode_sr",
    "identification",
    "add_documents",
    "skip_payment",
    "partial_settlement",
    "commercial_loan_extension",
    "retail_loan_amendment",
    "add_guarantor_sr",
    "block_unblock_service",
    "full_insurance_renew",
    "insurance_cancel",
    "delete_guarantor_sr",
    "total_loss",
    "charges_waiver",
    "deceased_fraud_sr",
    "legal_action_sr",
    "cancel_deal_request",
]

def create_sr_excel(name, folder="."):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Customer ID"
    ws["A2"] = "200125"
    filepath = os.path.join(folder, f"{name}.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")

if __name__ == "__main__":
    for sr in SERVICE_REQUESTS:
        create_sr_excel(sr)
    print(f"\nDone – {len(SERVICE_REQUESTS)} Excel files created.")
