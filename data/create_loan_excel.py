import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'LoanExtension'

hf = Font(bold=True, color='FFFFFF', size=11)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sf = Font(bold=True, color='FFFFFF', size=12)
sfill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
bfill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ca = Alignment(horizontal='center')

# Standard headers (common fields)
std_hdrs = ['Customer ID', 'Reschedulement_Type', 'New_Tenure',
            'New_Installment_Amount', 'Borrower_Status', 'Service_Charges_Waiver',
            'Payment_Mode', 'Remarks']

# Balloon headers (only balloon specific)
bal_hdrs = ['Balloon_Type', 'Balloon_Frequency', 'Balloon_Amount', 'Balloon_Installments']

# === STANDARD SECTION ===
for col in range(1, len(std_hdrs)+1):
    c = ws.cell(row=1, column=col)
    c.fill = sfill; c.font = sf; c.alignment = ca
ws['A1'] = '=== STANDARD SECTION ==='

for i, h in enumerate(std_hdrs):
    c = ws.cell(row=2, column=i+1, value=h)
    c.font = hf; c.fill = hfill; c.border = tb; c.alignment = ca

std_data = ['200125', 'Tenure', '48', '', 'Non Default', '0', 'Cash-Cheque', 'Standard loan extension']
for i, v in enumerate(std_data):
    c = ws.cell(row=3, column=i+1, value=v)
    c.border = tb; c.alignment = ca

# === BALLOON SECTION (Row 9) ===
for col in range(1, len(bal_hdrs)+1):
    c = ws.cell(row=9, column=col)
    c.fill = bfill; c.font = sf; c.alignment = ca
ws['A9'] = '=== BALLOON SECTION ==='

for i, h in enumerate(bal_hdrs):
    c = ws.cell(row=10, column=i+1, value=h)
    c.font = hf; c.fill = hfill; c.border = tb; c.alignment = ca

bal_data = ['Balloon', 'QUARTERLY', '1000', '4']
for i, v in enumerate(bal_data):
    c = ws.cell(row=11, column=i+1, value=v)
    c.border = tb; c.alignment = ca

# === DROPDOWNS ===
# Standard dropdowns
dv_resched = DataValidation(type='list', formula1='"Tenure,Installment"', allow_blank=True)
dv_resched.showDropDown = False
ws.add_data_validation(dv_resched)
dv_resched.add('B3')

dv_borrower = DataValidation(type='list', formula1='"Default,Non Default"', allow_blank=True)
dv_borrower.showDropDown = False
ws.add_data_validation(dv_borrower)
dv_borrower.add('E3')

dv_payment = DataValidation(type='list', formula1='"KNet-Kiosk,KNet-Web,Cash-Cheque"', allow_blank=True)
dv_payment.showDropDown = False
ws.add_data_validation(dv_payment)
dv_payment.add('G3')

# Balloon dropdowns
dv_btype = DataValidation(type='list', formula1='"Balloon,Regular EMI"', allow_blank=True)
dv_btype.showDropDown = False
ws.add_data_validation(dv_btype)
dv_btype.add('A11')

dv_bfreq = DataValidation(type='list', formula1='"MONTHLY,QUARTERLY,Half Yearly,Yearly"', allow_blank=True)
dv_bfreq.showDropDown = False
ws.add_data_validation(dv_bfreq)
dv_bfreq.add('B11')

# Column widths
for c in 'ABCDEFGH':
    ws.column_dimensions[c].width = 24

wb.save('commercial_loan_extension.xlsx')
print('DONE')
