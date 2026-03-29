from openpyxl import Workbook

# Create customer_search.xlsx
wb_customer = Workbook()
ws_customer = wb_customer.active
ws_customer['A1'] = 'Customer ID'
ws_customer['A2'] = '200125'
wb_customer.save('customer_search.xlsx')

# Create change_address.xlsx
wb_address = Workbook()
ws_address = wb_address.active

# Headers in row 1
headers = ['Action', 'Customer_Type', 'Status', 'Living_Since', 'Area', 'District', 'Block', 'Street', 'Building', 'Unit_Type', 'Unit_No', 'Floor']
for i, header in enumerate(headers, 1):
    ws_address.cell(row=1, column=i, value=header)

# RESIDENTIAL section header in row 2
ws_address['A2'] = 'RESIDENTIAL'

# RESIDENTIAL data in row 3
residential_data = ['Edit', 'Main Applicant', 'Active', '01/01/2020', 'SALMIYA', 'HAWALLI', '1', '10', 'Building A', 'Apartment', '101', '1']
for i, value in enumerate(residential_data, 1):
    ws_address.cell(row=3, column=i, value=value)

# PERMANENT section header in row 6
ws_address['A6'] = 'PERMANENT'

# PERMANENT data in row 7
permanent_data = ['Create', 'Main Applicant', 'Active', '01/01/2019', 'JABRIYA', 'HAWALLI', '2', '20', 'Building B', 'Villa', '201', '2']
for i, value in enumerate(permanent_data, 1):
    ws_address.cell(row=7, column=i, value=value)

# BUSINESS section header in row 10
ws_address['A10'] = 'BUSINESS'

# BUSINESS data in row 11
business_data = ['Skip', 'Main Applicant', 'Active', '01/01/2021', 'AHMADI', 'AHMADI', '3', '30', 'Building C', 'Apartment', '301', '3']
for i, value in enumerate(business_data, 1):
    ws_address.cell(row=11, column=i, value=value)

wb_address.save('change_address.xlsx')
print("Excel files created successfully!")