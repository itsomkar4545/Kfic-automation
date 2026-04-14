"""Create KNET Payment Test Cases Excel - Sheet1: Happy Path, Sheet2: Test Scenarios"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ── STYLES ──
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
edge_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin_border

def write_rows(ws, rows, start_row=2):
    for i, row in enumerate(rows):
        r = start_row + i
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = wrap
            cell.border = thin_border

# ════════════════════════════════════════════
# SHEET 1: Test Cases (Happy Path)
# ════════════════════════════════════════════
ws1 = wb.active
ws1.title = "KNET_Test_Cases"

headers1 = ["Step No", "Module", "Test Step Description", "Test Data", "Expected Result", "Actual Result", "Status (Pass/Fail)", "Remarks"]
style_header(ws1, headers1)

steps = [
    ["1", "Service Request", "Maker logs into LOS application", "Login ID: 20, Password: abcd@1234", "User is logged in successfully and dashboard is displayed"],
    ["2", "Service Request", "Navigate to Service Request module from menu", "Menu > Service > Service Request > Select SR type", "Service Request page is opened"],
    ["3", "Service Request", "Search customer by CIF number", "Customer ID: 200125", "Customer details are fetched and displayed"],
    ["4", "Service Request", "Select customer account from the list", "Select account radio button and click Proceed", "Account is selected and SR form is loaded"],
    ["5", "Service Request", "Fill all required Service Request details", "Fill mandatory fields as per SR type", "All fields are filled successfully without any error"],
    ["6", "Service Charges", "Check if service charges are applicable for this SR", "System auto-checks charges applicability", "Service charges amount is displayed if applicable"],
    ["7", "Payment Mode", "Select Payment Mode as 'KNet-Web'", "Payment Mode: KNet-Web", "KNet-Web is selected as payment mode"],
    ["8", "Service Request", "Upload required documents", "Upload document from local folder", "Document is uploaded successfully"],
    ["9", "Service Request", "Enter remarks and save the SR form", "Remarks: KNET payment for service charges", "SR form is saved successfully"],
    ["10", "Service Request", "Click 'Send for Verification' and assign to Manager", "Assign User: Manager user, Remarks: Send for verification", "SR is sent for verification to assigned manager"],
    ["11", "Service Request", "Maker logs out from application", "Click Logout", "User is logged out successfully"],
    ["12", "Approval", "Manager logs into LOS application", "Login ID: Manager user, Password: abcd@1234", "Manager is logged in and dashboard is displayed"],
    ["13", "Approval", "Manager opens Service Request Inbox", "Click Service Request Inbox icon", "Inbox is displayed with pending SR list"],
    ["14", "Approval", "Manager opens the SR raised by Maker (search by CIF)", "CIF: 200125", "SR details are displayed for review"],
    ["15", "Approval", "Manager approves the Service Request", "Click Approve button", "SR is approved successfully and KNET payment link is triggered"],
    ["16", "Approval", "Manager logs out from application", "Click Logout", "Manager is logged out successfully"],
    ["17", "KNET Link", "Customer receives KNET payment link on registered mobile number", "SMS sent to registered mobile number", "Customer receives SMS with KNET payment link"],
    ["18", "KNET Portal", "Customer opens the KNET payment link from SMS", "Click on payment link URL", "KNET payment portal login page is displayed"],
    ["19", "KNET Portal", "Customer enters registered mobile number on KNET portal", "Mobile No: Registered mobile number", "Mobile number is accepted and OTP is sent"],
    ["20", "KNET Portal", "Customer enters OTP received on mobile", "OTP: Received OTP", "OTP is verified successfully"],
    ["21", "KNET Portal", "KNET Home Page is displayed with service charges details", "Charges amount is shown for the SR", "Home page displays correct service charges amount to be paid"],
    ["22", "KNET Portal", "Customer clicks 'Pay Now' button", "Click Pay Now", "Payment details page is displayed / redirected"],
    ["23", "KNET Portal", "Customer selects Bank Name from dropdown", "Bank Name: e.g. NBK / KFH / CBK", "Bank is selected successfully"],
    ["24", "KNET Portal", "Customer enters Card Number", "Card No: Valid debit card number", "Card number is accepted"],
    ["25", "KNET Portal", "Customer enters Card Expiry Date", "Expiry: MM/YY (valid future date)", "Expiry date is accepted"],
    ["26", "KNET Portal", "Customer enters Card PIN", "PIN: Valid 4-digit PIN", "PIN is accepted"],
    ["27", "KNET Portal", "Customer clicks Submit button for payment", "Click Submit", "Payment confirmation popup/page is displayed"],
    ["28", "KNET Portal", "Customer clicks Confirm on final payment confirmation", "Click Confirm Payment", "Payment is processed successfully"],
    ["29", "KNET Portal", "Payment success page is displayed with Transaction/Reference ID", "Transaction ID / Reference ID is generated", "Success message with Transaction ID and Reference ID is displayed"],
    ["30", "KNET Portal", "Customer clicks Download/Print to get payment receipt", "Click Download Receipt", "Payment receipt is downloaded/printed with all transaction details"],
    ["31", "LMS Update", "Payment status is updated in LMS system", "Transaction ID mapped to SR", "LMS reflects payment status as 'Paid' and accounting entries are posted"],
    ["32", "Service Request", "Payment status (Success/Fail) is updated in Service Request", "SR status updated based on payment result", "SR shows payment status as 'Success'"],
    ["33", "Service Request", "Maker logs in and navigates to SR Inbox", "Login ID: 20, Password: abcd@1234", "SR is visible in inbox with updated payment status"],
    ["34", "Service Request", "Maker closes the Service Request", "Click Close button", "SR is closed successfully and flow is completed"],
    ["35", "Service Request", "Maker logs out from application", "Click Logout", "User is logged out. End-to-end KNET payment flow completed."],
]

for i, row in enumerate(steps):
    r = i + 2
    for col, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=col, value=val)
        cell.alignment = wrap
        cell.border = thin_border
    for col in [6, 7, 8]:
        cell = ws1.cell(row=r, column=col, value="")
        cell.alignment = wrap
        cell.border = thin_border

widths1 = [10, 18, 55, 45, 55, 30, 18, 25]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{len(steps)+1}"

# ════════════════════════════════════════════
# SHEET 2: Test Scenarios
# ════════════════════════════════════════════
ws2 = wb.create_sheet("KNET_Test_Scenarios")

headers2 = [
    "Scenario ID", "Scenario Type", "Module", "Test Scenario",
    "Pre-Condition", "Test Steps", "Test Data",
    "Expected Result", "Priority", "Status", "Remarks"
]
style_header(ws2, headers2)

type_fill_map = {"Positive": pos_fill, "Negative": neg_fill, "Edge Case": edge_fill}

scenarios = [
    # ── POSITIVE SCENARIOS ──
    ["TC_KNET_001", "Positive", "Service Request", "Successful SR creation with KNET payment mode",
     "User is logged in, customer exists", "Login > Navigate SR > Search CIF > Select Account > Fill SR > Select KNet-Web > Save",
     "Valid CIF, KNet-Web", "SR is created and saved successfully with KNET payment mode", "High"],
    ["TC_KNET_002", "Positive", "Approval", "Manager approves SR and KNET link is sent",
     "SR is submitted for verification", "Manager Login > Open Inbox > Open SR > Click Approve",
     "Valid SR with KNET mode", "SR approved, KNET payment link SMS sent to customer", "High"],
    ["TC_KNET_003", "Positive", "KNET Portal", "Customer login with registered mobile and OTP verification",
     "Customer received KNET link via SMS", "Open link > Enter mobile > Enter OTP",
     "Valid mobile, valid OTP", "OTP verified, KNET home page displayed with charges", "High"],
    ["TC_KNET_004", "Positive", "KNET Portal", "Successful KNET payment with valid card details",
     "Customer is on KNET payment page", "Select Bank > Enter Card No > Expiry > PIN > Submit > Confirm",
     "Valid bank, card, expiry, PIN", "Payment successful, Transaction ID generated", "High"],
    ["TC_KNET_005", "Positive", "KNET Portal", "Download/Print payment receipt after successful payment",
     "Payment is completed successfully", "Click Download/Print Receipt",
     "Transaction ID available", "Receipt downloaded with correct transaction details", "High"],
    ["TC_KNET_006", "Positive", "LMS", "LMS updated with payment status and accounting entries",
     "KNET payment completed", "Check LMS for payment status and accounting entries",
     "Transaction ID from KNET", "LMS shows Paid status, accounting entries posted correctly", "High"],
    ["TC_KNET_007", "Positive", "Service Request", "SR status updated to Success after payment",
     "KNET payment completed", "Check SR status in application",
     "SR linked to KNET transaction", "SR payment status shows 'Success'", "High"],
    ["TC_KNET_008", "Positive", "Service Request", "Maker closes SR after successful payment",
     "Payment done, SR status is Success", "Maker Login > Open Inbox > Open SR > Click Close",
     "Valid SR with payment done", "SR closed successfully, flow completed", "High"],

    # ── NEGATIVE SCENARIOS ──
    ["TC_KNET_009", "Negative", "KNET Portal", "Payment with invalid/wrong Card Number",
     "Customer is on payment page", "Enter wrong card number > Submit",
     "Invalid card number", "Error message displayed, payment is rejected", "High"],
    ["TC_KNET_010", "Negative", "KNET Portal", "Payment with expired Card",
     "Customer is on payment page", "Enter expired card expiry date > Submit",
     "Expiry: Past date (e.g. 01/22)", "Error message: Card expired, payment rejected", "High"],
    ["TC_KNET_011", "Negative", "KNET Portal", "Payment with wrong PIN",
     "Customer is on payment page", "Enter wrong PIN > Submit",
     "Wrong 4-digit PIN", "Error message: Invalid PIN, payment failed", "High"],
    ["TC_KNET_012", "Negative", "KNET Portal", "Payment with insufficient balance",
     "Customer is on payment page", "Enter valid card with low balance > Submit > Confirm",
     "Card with insufficient funds", "Error: Insufficient balance, payment declined", "High"],
    ["TC_KNET_013", "Negative", "KNET Portal", "Login with unregistered mobile number",
     "Customer has KNET link", "Open link > Enter unregistered mobile",
     "Unregistered mobile number", "Error: Mobile number not found / not registered", "Medium"],
    ["TC_KNET_014", "Negative", "KNET Portal", "Enter wrong/invalid OTP",
     "OTP sent to customer mobile", "Enter wrong OTP",
     "Invalid OTP", "Error: Invalid OTP, verification failed", "High"],
    ["TC_KNET_015", "Negative", "KNET Portal", "OTP expired - enter OTP after timeout",
     "OTP sent to customer mobile", "Wait for OTP to expire > Enter expired OTP",
     "Expired OTP", "Error: OTP expired, ask to resend", "Medium"],
    ["TC_KNET_016", "Negative", "KNET Portal", "Cancel payment on confirmation page",
     "Customer is on payment confirmation popup", "Click Cancel instead of Confirm",
     "N/A", "Payment is cancelled, user redirected back, no transaction created", "Medium"],
    ["TC_KNET_017", "Negative", "KNET Portal", "Submit payment with empty card details",
     "Customer is on payment page", "Leave all card fields empty > Click Submit",
     "Empty fields", "Validation error: All fields are mandatory", "Medium"],
    ["TC_KNET_018", "Negative", "KNET Portal", "Submit with partial card details (missing PIN)",
     "Customer is on payment page", "Enter bank, card, expiry but no PIN > Submit",
     "PIN field empty", "Validation error: PIN is required", "Medium"],
    ["TC_KNET_019", "Negative", "Service Request", "SR with no service charges - KNET option should not appear",
     "SR type has no charges", "Create SR where charges = 0",
     "SR with zero charges", "Payment mode selection is not displayed / disabled", "Medium"],

    # ── EDGE CASE SCENARIOS ──
    ["TC_KNET_020", "Edge Case", "KNET Portal", "Session timeout during payment process",
     "Customer is on payment page", "Wait for session to expire > Try to submit",
     "Session expired", "Session timeout error, user redirected to login page", "Medium"],
    ["TC_KNET_021", "Edge Case", "KNET Portal", "Network disconnection during payment",
     "Customer clicked Confirm", "Disconnect network during payment processing",
     "No network", "Appropriate error message, no duplicate transaction", "High"],
    ["TC_KNET_022", "Edge Case", "KNET Portal", "Double click on Pay Now / Confirm button",
     "Customer is on payment page", "Double click Pay Now or Confirm rapidly",
     "N/A", "Only one transaction is created, no duplicate payment", "High"],
    ["TC_KNET_023", "Edge Case", "KNET Portal", "Browser back button after successful payment",
     "Payment completed successfully", "Click browser back button",
     "N/A", "User should not be able to re-submit payment, appropriate message shown", "Medium"],
    ["TC_KNET_024", "Edge Case", "KNET Link", "Open expired KNET payment link",
     "Link was sent but time has passed", "Open link after expiry period",
     "Expired link URL", "Error: Payment link has expired", "Medium"],
    ["TC_KNET_025", "Edge Case", "KNET Link", "Open KNET payment link multiple times",
     "Link sent to customer", "Open same link in multiple tabs/browsers",
     "Same link URL", "Only one session is active, others show error", "Medium"],
    ["TC_KNET_026", "Edge Case", "KNET Portal", "Resend OTP and verify with new OTP",
     "First OTP expired or not received", "Click Resend OTP > Enter new OTP",
     "New OTP", "New OTP is sent, old OTP invalidated, new OTP works", "Medium"],
    ["TC_KNET_027", "Edge Case", "KNET Portal", "Payment with special characters in card fields",
     "Customer is on payment page", "Enter special chars in card number / PIN",
     "Card: @#$%, PIN: !@#$", "Validation error: Only numeric values allowed", "Low"],
    ["TC_KNET_028", "Edge Case", "LMS", "Payment failed but LMS status check",
     "KNET payment failed", "Check LMS after failed payment",
     "Failed transaction", "LMS shows payment status as 'Failed', no accounting entries posted", "High"],
    ["TC_KNET_029", "Edge Case", "Service Request", "SR status after payment failure",
     "KNET payment failed", "Check SR status after failed payment",
     "Failed payment SR", "SR shows payment status as 'Failed', SR remains open for retry", "High"],
    ["TC_KNET_030", "Edge Case", "KNET Portal", "Multiple wrong PIN attempts - card block scenario",
     "Customer is on payment page", "Enter wrong PIN 3+ times",
     "Wrong PIN multiple times", "Card is blocked / error message after max attempts", "Medium"],
]

for i, row in enumerate(scenarios):
    r = i + 2
    for col, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=col, value=val)
        cell.alignment = wrap
        cell.border = thin_border
    # Color code by scenario type
    scenario_type = row[1]
    if scenario_type in type_fill_map:
        ws2.cell(row=r, column=2).fill = type_fill_map[scenario_type]
    # Empty Status and Remarks
    for col in [10, 11]:
        cell = ws2.cell(row=r, column=col, value="")
        cell.alignment = wrap
        cell.border = thin_border

widths2 = [14, 14, 18, 45, 35, 50, 30, 50, 10, 12, 20]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:K{len(scenarios)+1}"

# Save
import os
filepath = os.path.join(os.path.dirname(__file__), "knet_payment_testcases.xlsx")
wb.save(filepath)
print(f"Created: {filepath}")
